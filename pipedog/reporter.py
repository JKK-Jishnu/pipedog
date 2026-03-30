"""
reporter.py — HTML and Excel scan report generation.

Generates a self-contained, human-readable HTML report and an Excel (.xlsx)
report after every `pipedog scan`.

HTML reports have zero external dependencies at render time — all CSS is
inline, no JavaScript frameworks, no CDN requests.

Excel reports are formatted workbooks with 3 sheets:
    Summary   — overall result, scan metadata, check counts
    Results   — all checks as a filterable table, color-coded by status
    Profile   — column statistics from the scanned file

Reports are saved to:
    .pipedog/<profile>/reports/<stem>-<profile>-<YYYYMMDD-HHMMSS>.html
    .pipedog/<profile>/reports/<stem>-<profile>-<YYYYMMDD-HHMMSS>.xlsx

Key functions:
    generate_html_report()   Build the HTML string from scan results.
    save_report()            Write the HTML file and return its path.
    generate_excel_report()  Build a formatted openpyxl Workbook.
    save_excel_report()      Write the .xlsx file and return its path.
    open_last_report()       Open the most recent report in the system browser.
"""

from __future__ import annotations

import webbrowser
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from . import __version__
from .profiler import _pipedog_dir
from .schema import CheckResult, DataSchema

# ---------------------------------------------------------------------------
# CSS (inline, no external dependencies)
# ---------------------------------------------------------------------------

_CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    font-size: 14px;
    color: #1f2937;
    background: #f9fafb;
    padding: 24px;
}
h1 { font-size: 22px; font-weight: 700; margin-bottom: 4px; }
h2 { font-size: 15px; font-weight: 600; margin: 24px 0 10px; }
.meta { color: #6b7280; font-size: 12px; margin-bottom: 20px; }
.summary {
    border-radius: 8px;
    padding: 16px 20px;
    margin-bottom: 24px;
    border-left: 6px solid #ccc;
    background: #fff;
}
.summary.pass  { border-color: #16a34a; background: #f0fdf4; }
.summary.warn  { border-color: #d97706; background: #fffbeb; }
.summary.fail  { border-color: #dc2626; background: #fef2f2; }
.summary .status { font-size: 18px; font-weight: 700; margin-bottom: 6px; }
.summary.pass .status  { color: #16a34a; }
.summary.warn .status  { color: #d97706; }
.summary.fail .status  { color: #dc2626; }
.summary .counts { color: #6b7280; font-size: 13px; }
table {
    width: 100%;
    border-collapse: collapse;
    background: #fff;
    border-radius: 8px;
    overflow: hidden;
    box-shadow: 0 1px 3px rgba(0,0,0,0.08);
    margin-bottom: 8px;
}
th {
    background: #f3f4f6;
    padding: 10px 14px;
    text-align: left;
    font-size: 12px;
    font-weight: 600;
    color: #6b7280;
    text-transform: uppercase;
    letter-spacing: 0.05em;
}
td { padding: 10px 14px; border-top: 1px solid #f3f4f6; font-size: 13px; }
tr:hover td { background: #f9fafb; }
.badge {
    display: inline-block;
    padding: 2px 8px;
    border-radius: 12px;
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 0.03em;
}
.badge-pass { background: #dcfce7; color: #16a34a; }
.badge-fail { background: #fee2e2; color: #dc2626; }
.badge-warn { background: #fef3c7; color: #d97706; }
.badge-drift { background: #e0e7ff; color: #4338ca; }
.section-empty { color: #9ca3af; font-style: italic; padding: 8px 0; }
.report-path { font-size: 11px; color: #9ca3af; margin-top: 32px; }
"""

# ---------------------------------------------------------------------------
# HTML builders
# ---------------------------------------------------------------------------

def _esc(text: str) -> str:
    """HTML-escape a string to prevent XSS in report values."""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _badge(label: str, kind: str) -> str:
    return f'<span class="badge badge-{kind}">{_esc(label)}</span>'


def _results_table(results: list[CheckResult], include_status: bool = True) -> str:
    if not results:
        return '<p class="section-empty">None.</p>'
    rows = ""
    for r in results:
        if r.passed:
            status = _badge("PASS", "pass")
        elif r.severity == "warning":
            status = _badge("WARN", "warn")
        else:
            status = _badge("FAIL", "fail")

        col_display = r.column if r.column != "__row_count__" else "File (row count)"
        rows += (
            f"<tr>"
            f"<td>{_esc(col_display)}</td>"
            f"<td>{_esc(r.check_type)}</td>"
            f"<td>{status}</td>"
            f"<td>{_esc(r.detail)}</td>"
            f"</tr>"
        )
    return (
        "<table><thead><tr>"
        "<th>Column</th><th>Check Type</th><th>Status</th><th>Detail</th>"
        "</tr></thead><tbody>"
        + rows
        + "</tbody></table>"
    )


def _drift_table(drift_results: list[CheckResult]) -> str:
    if not drift_results:
        return '<p class="section-empty">No structural drift detected.</p>'
    rows = ""
    for r in drift_results:
        sev = _badge("WARN", "warn") if r.severity == "warning" else _badge("FAIL", "fail")
        rows += (
            f"<tr>"
            f"<td>{_esc(r.column)}</td>"
            f"<td>{_badge(r.check_type.replace('_', ' ').title(), 'drift')}</td>"
            f"<td>{sev}</td>"
            f"<td>{_esc(r.detail)}</td>"
            f"</tr>"
        )
    return (
        "<table><thead><tr>"
        "<th>Column</th><th>Type</th><th>Severity</th><th>Detail</th>"
        "</tr></thead><tbody>"
        + rows
        + "</tbody></table>"
    )


def _column_profile_table(schema: DataSchema) -> str:
    rows = ""
    for col in schema.columns:
        null_color = "#dc2626" if col.null_count > 0 else "#16a34a"
        null_cell = (
            f'<td style="color:{null_color}">'
            f"{col.null_count} ({col.null_pct}%)</td>"
        )
        min_str = f"{col.min_value:,.4g}" if col.min_value is not None else "-"
        max_str = f"{col.max_value:,.4g}" if col.max_value is not None else "-"
        std_str = f"{col.std_dev:,.4g}" if col.std_dev is not None else "-"
        sample_str = _esc(", ".join(str(v) for v in col.sample_values[:3]))
        unique_str = str(col.unique_count) if col.unique_count != -1 else "merged"
        rows += (
            f"<tr>"
            f"<td><strong>{_esc(col.name)}</strong></td>"
            f"<td>{_esc(col.dtype)}</td>"
            f"{null_cell}"
            f"<td>{unique_str}</td>"
            f"<td>{min_str}</td>"
            f"<td>{max_str}</td>"
            f"<td>{std_str}</td>"
            f"<td>{sample_str}</td>"
            f"</tr>"
        )
    return (
        "<table><thead><tr>"
        "<th>Column</th><th>Type</th><th>Nulls</th>"
        "<th>Unique</th><th>Min</th><th>Max</th><th>Std Dev</th>"
        "<th>Sample Values</th>"
        "</tr></thead><tbody>"
        + rows
        + "</tbody></table>"
    )


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_html_report(
    drift_results: list[CheckResult],
    check_results: list[CheckResult],
    current_schema: DataSchema,
    baseline_schema: DataSchema,
    profile: Optional[str],
    scanned_file: str,
) -> str:
    """
    Build a self-contained HTML scan report as a string.

    The report contains five sections:
        1. Summary header (pass/warn/fail banner with counts)
        2. Schema Drift (added/removed columns, type changes)
        3. Quality Check Results (all checks in one table)
        4. Column Profile (stats for every column in the scanned file)

    All styling is inline CSS — no external resources required.

    Args:
        drift_results:   CheckResult list from scanner.detect_drift().
        check_results:   CheckResult list from scanner.run_quality_checks().
        current_schema:  DataSchema profiled from the current file.
        baseline_schema: DataSchema loaded from the saved snapshot.
        profile:         Profile name, or None for the default profile.
        scanned_file:    Path to the file that was scanned.

    Returns:
        A complete HTML document as a string.
    """
    all_results = drift_results + check_results
    failures = [r for r in all_results if not r.passed and r.severity == "error"]
    warnings_list = [r for r in all_results if not r.passed and r.severity == "warning"]
    passed_list = [r for r in all_results if r.passed]

    overall_pass = len(failures) == 0

    if overall_pass and not warnings_list:
        status_class = "pass"
        status_label = "ALL CHECKS PASSED"
    elif overall_pass:
        status_class = "warn"
        status_label = "PASSED WITH WARNINGS"
    else:
        status_class = "fail"
        status_label = "CHECKS FAILED"

    profile_label = profile or "default"
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    counts_str = (
        f"{current_schema.row_count:,} rows &nbsp;·&nbsp; "
        f"{current_schema.column_count} columns &nbsp;·&nbsp; "
        f"{len(passed_list)} passed &nbsp;·&nbsp; "
        f"{len(warnings_list)} warnings &nbsp;·&nbsp; "
        f"{len(failures)} failed"
    )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Pipedog Scan — {_esc(Path(scanned_file).name)}</title>
<style>{_CSS}</style>
</head>
<body>

<h1>Pipedog Scan Report</h1>
<p class="meta">
  File: <strong>{_esc(scanned_file)}</strong> &nbsp;·&nbsp;
  Profile: <strong>{_esc(profile_label)}</strong> &nbsp;·&nbsp;
  {_esc(timestamp)}
</p>

<div class="summary {status_class}">
  <div class="status">{status_label}</div>
  <div class="counts">{counts_str}</div>
</div>

<h2>Schema Drift</h2>
{_drift_table(drift_results)}

<h2>Quality Checks</h2>
{_results_table(all_results)}

<h2>Column Profile (Current File)</h2>
{_column_profile_table(current_schema)}

<p class="report-path">
  Baseline snapshot: {_esc(baseline_schema.captured_at[:19].replace("T", " "))} UTC
  &nbsp;&middot;&nbsp; Generated by Pipedog v{__version__}
</p>

</body>
</html>"""

    return html


def save_report(
    html: str,
    profile: Optional[str],
    scanned_file: str,
) -> Path:
    """
    Write an HTML report string to the reports directory and return its path.

    File naming pattern:
        .pipedog/<profile>/reports/<stem>-<profile>-<YYYYMMDD-HHMMSS>.html

    Creates the reports/ directory if it does not exist.

    Args:
        html:         The HTML string from generate_html_report().
        profile:      Profile name, or None for the default profile.
        scanned_file: Path to the scanned file (used to derive the stem).

    Returns:
        The absolute Path of the written HTML file.
    """
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    stem = Path(scanned_file).stem
    profile_label = profile or "default"
    filename = f"{stem}-{profile_label}-{timestamp}.html"

    reports_dir = _pipedog_dir(profile) / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)

    report_path = reports_dir / filename
    report_path.write_text(html, encoding="utf-8")
    return report_path.resolve()


def generate_excel_report(
    drift_results: list[CheckResult],
    check_results: list[CheckResult],
    current_schema: DataSchema,
    baseline_schema: DataSchema,
    profile: Optional[str],
    scanned_file: str,
) -> Workbook:
    """
    Build a formatted Excel workbook from scan results.

    Sheets:
        Summary  — overall pass/fail banner, scan metadata, check counts.
        Results  — all drift + quality check results as a filterable table,
                   rows color-coded green/red/yellow by status.
        Profile  — per-column statistics from the scanned file, null counts
                   highlighted red when non-zero.

    Color scheme matches Pipedog terminal output:
        Pass    — green  (#16a34a fill, #dcfce7 row tint)
        Fail    — red    (#dc2626 fill, #fef2f2 row tint)
        Warning — amber  (#d97706 fill, #fef3c7 row tint)

    Args:
        drift_results:   CheckResult list from scanner.detect_drift().
        check_results:   CheckResult list from scanner.run_quality_checks().
        current_schema:  DataSchema profiled from the current file.
        baseline_schema: DataSchema loaded from the saved snapshot.
        profile:         Profile name, or None for the default profile.
        scanned_file:    Path to the file that was scanned.

    Returns:
        An openpyxl Workbook ready to be saved with save_excel_report().
    """
    all_results = drift_results + check_results
    failures = [r for r in all_results if not r.passed and r.severity == "error"]
    warnings_list = [r for r in all_results if not r.passed and r.severity == "warning"]
    passed_list = [r for r in all_results if r.passed]
    overall_pass = len(failures) == 0

    profile_label = profile or "default"
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    # --- Color palette ---
    GREEN_DARK  = PatternFill("solid", fgColor="16a34a")
    GREEN_LIGHT = PatternFill("solid", fgColor="dcfce7")
    RED_DARK    = PatternFill("solid", fgColor="dc2626")
    RED_LIGHT   = PatternFill("solid", fgColor="fef2f2")
    AMBER_DARK  = PatternFill("solid", fgColor="d97706")
    AMBER_LIGHT = PatternFill("solid", fgColor="fef3c7")
    GREY_HEADER = PatternFill("solid", fgColor="f3f4f6")
    WHITE       = PatternFill("solid", fgColor="FFFFFF")

    BOLD       = Font(bold=True)
    BOLD_WHITE = Font(bold=True, color="FFFFFF")
    HEADER_FONT = Font(bold=True, color="374151")

    thin = Side(style="thin", color="e5e7eb")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center")
    WRAP   = Alignment(wrap_text=True, vertical="top")

    def _header_row(ws, headers: list[str], row: int = 1) -> None:
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = GREY_HEADER
            cell.border = BORDER
            cell.alignment = CENTER

    def _set_col_widths(ws, widths: list[int]) -> None:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = Workbook()

    # ---------------------------------------------------------------
    # Sheet 1 — Summary
    # ---------------------------------------------------------------
    ws_sum = wb.active
    ws_sum.title = "Summary"

    if overall_pass and not warnings_list:
        status_text = "ALL CHECKS PASSED"
        status_fill = GREEN_DARK
    elif overall_pass:
        status_text = "PASSED WITH WARNINGS"
        status_fill = AMBER_DARK
    else:
        status_text = "CHECKS FAILED"
        status_fill = RED_DARK

    ws_sum.merge_cells("A1:B1")
    cell = ws_sum["A1"]
    cell.value = status_text
    cell.font = Font(bold=True, size=16, color="FFFFFF")
    cell.fill = status_fill
    cell.alignment = CENTER
    ws_sum.row_dimensions[1].height = 36

    ws_sum.append([])  # spacer

    meta = [
        ("File scanned",      str(Path(scanned_file).resolve())),
        ("Profile",           profile_label),
        ("Scan time",         timestamp),
        ("Baseline captured", baseline_schema.captured_at[:19].replace("T", " ") + " UTC"),
        ("Pipedog version",   f"v{__version__}"),
    ]
    for label, value in meta:
        row = ws_sum.max_row + 1
        ws_sum.cell(row=row, column=1, value=label).font = BOLD
        ws_sum.cell(row=row, column=2, value=value)

    ws_sum.append([])  # spacer

    count_headers = ["Total Checks", "Passed", "Warnings", "Failed", "Rows", "Columns"]
    count_values  = [
        len(all_results), len(passed_list), len(warnings_list),
        len(failures), current_schema.row_count, current_schema.column_count,
    ]
    hrow = ws_sum.max_row + 1
    for col, h in enumerate(count_headers, start=1):
        c = ws_sum.cell(row=hrow, column=col, value=h)
        c.font = HEADER_FONT; c.fill = GREY_HEADER; c.alignment = CENTER; c.border = BORDER
    vrow = hrow + 1
    fills = [WHITE, GREEN_LIGHT, AMBER_LIGHT, RED_LIGHT, WHITE, WHITE]
    for col, (val, fill) in enumerate(zip(count_values, fills), start=1):
        c = ws_sum.cell(row=vrow, column=col, value=val)
        c.font = BOLD; c.fill = fill; c.alignment = CENTER; c.border = BORDER

    _set_col_widths(ws_sum, [22, 55, 16, 12, 12, 12])

    # ---------------------------------------------------------------
    # Sheet 2 — Results
    # ---------------------------------------------------------------
    ws_res = wb.create_sheet("Results")
    headers = ["Column", "Check Type", "Status", "Severity", "Detail"]
    _header_row(ws_res, headers)
    ws_res.auto_filter.ref = "A1:E1"

    for r in all_results:
        if r.passed:
            label, row_fill = "PASS", GREEN_LIGHT
        elif r.severity == "warning":
            label, row_fill = "WARN", AMBER_LIGHT
        else:
            label, row_fill = "FAIL", RED_LIGHT

        col_display = r.column if r.column != "__row_count__" else "File (row count)"
        row_idx = ws_res.max_row + 1
        values = [col_display, r.check_type, label, r.severity, r.detail]
        for col, val in enumerate(values, start=1):
            c = ws_res.cell(row=row_idx, column=col, value=val)
            c.fill = row_fill
            c.border = BORDER
            c.alignment = WRAP

    _set_col_widths(ws_res, [22, 18, 10, 10, 70])
    ws_res.row_dimensions[1].height = 18

    # ---------------------------------------------------------------
    # Sheet 3 — Profile
    # ---------------------------------------------------------------
    ws_pro = wb.create_sheet("Profile")
    pro_headers = ["Column", "Type", "Nulls", "Null %", "Unique", "Min", "Max", "Std Dev", "Sample Values"]
    _header_row(ws_pro, pro_headers)

    for col in current_schema.columns:
        null_fill = RED_LIGHT if col.null_count > 0 else WHITE
        unique_str = str(col.unique_count) if col.unique_count != -1 else "merged"
        min_str = f"{col.min_value:,.4g}" if col.min_value is not None else "-"
        max_str = f"{col.max_value:,.4g}" if col.max_value is not None else "-"
        std_str = f"{col.std_dev:,.4g}" if col.std_dev is not None else "-"
        sample_str = ", ".join(str(v) for v in col.sample_values[:3])

        row_idx = ws_pro.max_row + 1
        values = [col.name, col.dtype, col.null_count, f"{col.null_pct}%",
                  unique_str, min_str, max_str, std_str, sample_str]
        for c_idx, val in enumerate(values, start=1):
            c = ws_pro.cell(row=row_idx, column=c_idx, value=val)
            c.border = BORDER
            c.alignment = Alignment(vertical="center")
            if c_idx in (3, 4):  # null columns
                c.fill = null_fill
                if col.null_count > 0:
                    c.font = Font(color="dc2626", bold=True)

    _set_col_widths(ws_pro, [20, 12, 8, 9, 9, 12, 12, 10, 35])

    return wb


def save_excel_report(
    wb: Workbook,
    profile: Optional[str],
    scanned_file: str,
) -> Path:
    """
    Write an openpyxl Workbook to the reports directory and return its path.

    File naming pattern:
        .pipedog/<profile>/reports/<stem>-<profile>-<YYYYMMDD-HHMMSS>.xlsx

    Creates the reports/ directory if it does not exist.

    Args:
        wb:           The Workbook from generate_excel_report().
        profile:      Profile name, or None for the default profile.
        scanned_file: Path to the scanned file (used to derive the stem).

    Returns:
        The absolute Path of the written .xlsx file.
    """
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    stem = Path(scanned_file).stem
    profile_label = profile or "default"
    filename = f"{stem}-{profile_label}-{timestamp}.xlsx"

    reports_dir = _pipedog_dir(profile) / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)

    report_path = reports_dir / filename
    wb.save(str(report_path))
    return report_path.resolve()


def open_last_report(profile: Optional[str]) -> None:
    """
    Open the most recently modified HTML report in the system default browser.

    Args:
        profile: Profile name, or None for the default profile.

    Raises:
        FileNotFoundError: If no reports exist for the given profile.
    """
    reports_dir = _pipedog_dir(profile) / "reports"
    if not reports_dir.exists():
        raise FileNotFoundError(
            f"No reports directory found at {reports_dir}. "
            "Run `pipedog scan` first."
        )
    reports = sorted(
        reports_dir.glob("*.html"),
        key=lambda p: p.stat().st_mtime,
    )
    if not reports:
        raise FileNotFoundError(
            f"No HTML reports found in {reports_dir}. "
            "Run `pipedog scan` first."
        )
    webbrowser.open(reports[-1].resolve().as_uri())
